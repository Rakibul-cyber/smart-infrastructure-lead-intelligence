from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .crawler import CrawledWebsite
from .scorer import LeadScore
from .signal_detector import DetectedSignals, SignalEvidence


logger = logging.getLogger(__name__)

LEAD_HEADERS = [
    "Organisation Name",
    "Organisation Type",
    "City",
    "State",
    "Website",
    "Visited Pages",
    "Emails",
    "Phone Numbers",
    "Contact Links",
    "Document Links",
    "Street Lighting",
    "Smart City",
    "Energy Efficiency",
    "Climate Action",
    "Infrastructure Modernisation",
    "Procurement",
    "Municipal Utility",
    "Matched Keywords",
    "Evidence Count",
    "Lead Score",
    "Priority",
    "Score Summary",
    "Score Breakdown",
    "Last Checked",
]

EVIDENCE_HEADERS = [
    "Organisation Name",
    "Website",
    "Category",
    "Keyword",
    "Excerpt",
    "Source URL",
]

PRIORITY_FILLS = {
    "High": PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    ),
    "Medium": PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    ),
    "Low": PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    ),
}

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="E7EEF7",
)


@dataclass
class LeadRecord:
    """Export-ready lead information for one analysed organisation."""

    organisation_name: str
    organisation_type: str
    city: str
    state: str
    website: str
    visited_pages: int
    emails: list[str]
    phone_numbers: list[str]
    contact_links: list[str]
    street_lighting: bool
    smart_city: bool
    energy_efficiency: bool
    climate_action: bool
    infrastructure_modernisation: bool
    procurement: bool
    municipal_utility: bool
    matched_keywords: list[str]
    evidence_count: int
    lead_score: int
    priority: str
    score_summary: str
    score_breakdown: list[str]
    last_checked: str
    document_links: list[str] = field(default_factory=list)


def build_lead_record(
    organisation_name: str,
    organisation_type: str,
    city: str,
    state: str,
    website: str,
    crawler_result: CrawledWebsite,
    signals: DetectedSignals,
    lead_score: LeadScore,
    last_checked: str,
) -> LeadRecord:
    """
    Map crawler, signal, and scoring results into an export-ready record.

    This function preserves list values and does not write any files.
    """

    return LeadRecord(
        organisation_name=organisation_name,
        organisation_type=organisation_type,
        city=city,
        state=state,
        website=website,
        visited_pages=len(crawler_result.visited_urls),
        emails=crawler_result.emails,
        phone_numbers=crawler_result.phone_numbers,
        contact_links=crawler_result.contact_links,
        document_links=crawler_result.document_links,
        street_lighting=signals.street_lighting,
        smart_city=signals.smart_city,
        energy_efficiency=signals.energy_efficiency,
        climate_action=signals.climate_action,
        infrastructure_modernisation=signals.infrastructure_modernisation,
        procurement=signals.procurement,
        municipal_utility=signals.municipal_utility,
        matched_keywords=signals.matched_keywords,
        evidence_count=len(signals.evidence),
        lead_score=lead_score.total_score,
        priority=lead_score.priority,
        score_summary=lead_score.summary,
        score_breakdown=[
            _format_score_breakdown_item(item.points, item.criterion, item.reason)
            for item in lead_score.breakdown
        ],
        last_checked=last_checked,
    )


def export_leads_to_excel(
    records: list[LeadRecord],
    evidence_by_website: dict[str, list[SignalEvidence]],
    output_path: str | Path,
) -> Path:
    """
    Export analysed lead records and evidence to a styled Excel workbook.

    Returns the created output path. The exact requested file may be
    overwritten, and parent directories are created automatically.
    """

    if not records:
        raise ValueError("records must not be empty")

    path = Path(output_path).expanduser()
    evidence_count = sum(
        len(evidence_items)
        for evidence_items in evidence_by_website.values()
    )
    logger.info("Excel export started")
    logger.debug(
        "Excel export details: records=%d evidence_items=%d output_path=%s",
        len(records),
        evidence_count,
        path,
    )
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    record_by_website = {
        record.website: record
        for record in records
    }
    unknown_evidence_websites = set(evidence_by_website) - set(record_by_website)

    if unknown_evidence_websites:
        unknown_websites = ", ".join(sorted(unknown_evidence_websites))
        raise ValueError(
            "evidence_by_website contains websites without records: "
            f"{unknown_websites}"
        )

    workbook = Workbook()
    all_leads_sheet = workbook.active
    all_leads_sheet.title = "All Leads"

    _write_lead_sheet(
        worksheet=all_leads_sheet,
        records=records,
    )
    _write_lead_sheet(
        worksheet=workbook.create_sheet("High Priority"),
        records=[
            record
            for record in records
            if record.priority == "High"
        ],
    )
    _write_evidence_sheet(
        worksheet=workbook.create_sheet("Evidence"),
        records=records,
        evidence_by_website=evidence_by_website,
    )
    _write_run_summary_sheet(
        worksheet=workbook.create_sheet("Run Summary"),
        records=records,
        evidence_by_website=evidence_by_website,
    )

    workbook.save(path)
    workbook.close()
    logger.info("Workbook saved successfully: %s", path)
    logger.debug("Workbook sheets: %s", workbook.sheetnames)

    # Verify the workbook can be reopened after saving.
    loaded_workbook = load_workbook(path)
    loaded_workbook.close()

    return path.resolve()


def _write_lead_sheet(
    worksheet: Worksheet,
    records: list[LeadRecord],
) -> None:
    """Write a lead table to a worksheet."""

    worksheet.append(LEAD_HEADERS)

    for record in records:
        worksheet.append(_lead_record_to_row(record))

    _style_table_sheet(worksheet)
    _style_priority_cells(worksheet)
    _style_score_column(worksheet)


def _write_evidence_sheet(
    worksheet: Worksheet,
    records: list[LeadRecord],
    evidence_by_website: dict[str, list[SignalEvidence]],
) -> None:
    """Write one evidence row for each SignalEvidence object."""

    worksheet.append(EVIDENCE_HEADERS)

    record_by_website = {
        record.website: record
        for record in records
    }

    for record in records:
        for evidence_item in evidence_by_website.get(record.website, []):
            worksheet.append(
                [
                    record_by_website[record.website].organisation_name,
                    record.website,
                    evidence_item.category,
                    evidence_item.keyword,
                    evidence_item.excerpt,
                    evidence_item.source_url,
                ]
            )

    _style_table_sheet(worksheet)


def _write_run_summary_sheet(
    worksheet: Worksheet,
    records: list[LeadRecord],
    evidence_by_website: dict[str, list[SignalEvidence]],
) -> None:
    """Write aggregate export summary values."""

    exported_at = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")
    total_evidence_items = sum(
        len(evidence_items)
        for evidence_items in evidence_by_website.values()
    )
    summary_rows = [
        ("Total Leads", len(records)),
        ("High Priority Leads", _count_priority(records, "High")),
        ("Medium Priority Leads", _count_priority(records, "Medium")),
        ("Low Priority Leads", _count_priority(records, "Low")),
        (
            "Total Emails Found",
            sum(len(record.emails) for record in records),
        ),
        (
            "Total Phone Numbers Found",
            sum(len(record.phone_numbers) for record in records),
        ),
        (
            "Total Document Links Found",
            sum(len(record.document_links) for record in records),
        ),
        ("Total Evidence Items", total_evidence_items),
        (
            "Average Lead Score",
            mean(record.lead_score for record in records),
        ),
        ("Exported At", exported_at),
    ]

    for key, value in summary_rows:
        worksheet.append([key, value])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 24
    worksheet["B9"].number_format = "0.00"


def _lead_record_to_row(record: LeadRecord) -> list[Any]:
    """Convert a LeadRecord into the All Leads column order."""

    return [
        record.organisation_name,
        record.organisation_type,
        record.city,
        record.state,
        record.website,
        record.visited_pages,
        _join_list_values(record.emails),
        _join_list_values(record.phone_numbers),
        _join_list_values(record.contact_links),
        _join_list_values(record.document_links),
        _format_bool(record.street_lighting),
        _format_bool(record.smart_city),
        _format_bool(record.energy_efficiency),
        _format_bool(record.climate_action),
        _format_bool(record.infrastructure_modernisation),
        _format_bool(record.procurement),
        _format_bool(record.municipal_utility),
        _join_list_values(record.matched_keywords),
        record.evidence_count,
        record.lead_score,
        record.priority,
        record.score_summary,
        _join_list_values(record.score_breakdown),
        record.last_checked,
    ]


def _style_table_sheet(worksheet: Worksheet) -> None:
    """Apply neutral professional styling to a tabular worksheet."""

    if worksheet.max_row < 1:
        return

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _apply_column_widths(worksheet)


def _apply_column_widths(worksheet: Worksheet) -> None:
    """Set readable widths without overdesigning the workbook."""

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value))
            for cell in column_cells
            if cell.value is not None
        )
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            45,
        )


def _style_priority_cells(worksheet: Worksheet) -> None:
    """Apply priority-specific fills in the Priority column."""

    priority_column = LEAD_HEADERS.index("Priority") + 1

    for row_number in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(
            row=row_number,
            column=priority_column,
        )
        fill = PRIORITY_FILLS.get(str(cell.value))

        if fill is not None:
            cell.fill = fill


def _style_score_column(worksheet: Worksheet) -> None:
    """Apply integer formatting to lead score cells."""

    score_column = LEAD_HEADERS.index("Lead Score") + 1

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(
            row=row_number,
            column=score_column,
        ).number_format = "0"


def _format_score_breakdown_item(
    points: int,
    criterion: str,
    reason: str,
) -> str:
    """Format one score breakdown item for Excel export."""

    return (
        f"+{points} "
        f"{criterion.replace('_', ' ').capitalize()} — "
        f"{reason}"
    )


def _join_list_values(values: list[str]) -> str:
    """Join list values for one spreadsheet cell."""

    return "; ".join(values)


def _format_bool(value: bool) -> str:
    """Format booleans for spreadsheet readability."""

    if value:
        return "Yes"

    return "No"


def _count_priority(
    records: list[LeadRecord],
    priority: str,
) -> int:
    """Count records in one priority band."""

    return sum(
        1
        for record in records
        if record.priority == priority
    )
