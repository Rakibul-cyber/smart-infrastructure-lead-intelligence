from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.lead_intelligence.crawler import CrawledWebsite
from src.lead_intelligence.exporter import (
    LEAD_HEADERS,
    LeadRecord,
    build_lead_record,
    export_leads_to_excel,
)
from src.lead_intelligence.scorer import (
    LeadScore,
    ScoreBreakdownItem,
)
from src.lead_intelligence.signal_detector import (
    DetectedSignals,
    SignalEvidence,
)


def make_crawler_result() -> CrawledWebsite:
    """Build a deterministic crawler result for exporter tests."""

    return CrawledWebsite(
        start_url="https://example-city.de",
        visited_urls=[
            "https://example-city.de",
            "https://example-city.de/kontakt",
        ],
        failed_pages=[],
        emails=[
            "office@example-city.de",
            "energy@example-city.de",
        ],
        phone_numbers=["030 1234 5678"],
        contact_links=["https://example-city.de/kontakt"],
        page_results=[],
    )


def make_signals() -> DetectedSignals:
    """Build deterministic signal results for exporter tests."""

    return DetectedSignals(
        street_lighting=True,
        smart_city=True,
        energy_efficiency=False,
        climate_action=False,
        infrastructure_modernisation=True,
        procurement=False,
        municipal_utility=True,
        matched_keywords=[
            "smart city",
            "street lighting",
        ],
        evidence=[
            SignalEvidence(
                category="street_lighting",
                keyword="street lighting",
                excerpt="Fictional street lighting programme.",
                source_url="https://example-city.de",
            )
        ],
    )


def make_lead_score() -> LeadScore:
    """Build a deterministic score for exporter tests."""

    return LeadScore(
        total_score=55,
        priority="Medium",
        breakdown=[
            ScoreBreakdownItem(
                criterion="street_lighting",
                points=25,
                reason=(
                    "Street-lighting or public-lighting activity was detected."
                ),
            ),
            ScoreBreakdownItem(
                criterion="smart_city",
                points=15,
                reason=(
                    "A Smart City or municipal digitalisation initiative was detected."
                ),
            ),
        ],
        summary="Medium-priority lead with relevant signals.",
    )


def make_record(
    *,
    name: str = "Example City",
    website: str = "https://example-city.de",
    priority: str = "Medium",
    score: int = 55,
    emails: list[str] | None = None,
    phone_numbers: list[str] | None = None,
) -> LeadRecord:
    """Build a fictional LeadRecord."""

    return LeadRecord(
        organisation_name=name,
        organisation_type="Municipality",
        city="Example City",
        state="Example State",
        website=website,
        visited_pages=2,
        emails=emails
        if emails is not None
        else ["office@example-city.de"],
        phone_numbers=phone_numbers
        if phone_numbers is not None
        else ["030 1234 5678"],
        contact_links=["https://example-city.de/kontakt"],
        street_lighting=True,
        smart_city=priority != "Low",
        energy_efficiency=False,
        climate_action=False,
        infrastructure_modernisation=priority == "High",
        procurement=priority == "High",
        municipal_utility=True,
        matched_keywords=[
            "smart city",
            "street lighting",
        ],
        evidence_count=1,
        lead_score=score,
        priority=priority,
        score_summary=f"{priority}-priority fictional record.",
        score_breakdown=[
            "+25 Street lighting — Street-lighting or public-lighting activity was detected."
        ],
        last_checked="2026-07-30",
    )


def make_evidence(
    website: str = "https://example-city.de",
) -> SignalEvidence:
    """Build fictional evidence."""

    return SignalEvidence(
        category="street_lighting",
        keyword="street lighting",
        excerpt="Fictional street lighting evidence.",
        source_url=f"{website}/signals",
    )


def load_export(path: Path):
    """Open an exported workbook for assertions."""

    return load_workbook(path)


def test_build_lead_record_maps_fields_correctly() -> None:
    """Crawler, signal, and score fields should map into LeadRecord."""

    record = build_lead_record(
        organisation_name="Example City",
        organisation_type="Municipality",
        city="Example City",
        state="Example State",
        website="https://example-city.de",
        crawler_result=make_crawler_result(),
        signals=make_signals(),
        lead_score=make_lead_score(),
        last_checked="2026-07-30",
    )

    assert record.organisation_name == "Example City"
    assert record.visited_pages == 2
    assert record.emails == [
        "office@example-city.de",
        "energy@example-city.de",
    ]
    assert record.street_lighting is True
    assert record.energy_efficiency is False
    assert record.lead_score == 55
    assert record.priority == "Medium"
    assert record.evidence_count == 1


def test_score_breakdown_strings_are_created_correctly() -> None:
    """Score breakdown items should become readable export strings."""

    record = build_lead_record(
        organisation_name="Example City",
        organisation_type="Municipality",
        city="Example City",
        state="Example State",
        website="https://example-city.de",
        crawler_result=make_crawler_result(),
        signals=make_signals(),
        lead_score=make_lead_score(),
        last_checked="2026-07-30",
    )

    assert record.score_breakdown[0] == (
        "+25 Street lighting — "
        "Street-lighting or public-lighting activity was detected."
    )


def test_empty_record_list_raises_value_error(tmp_path: Path) -> None:
    """Exporting without records should fail clearly."""

    with pytest.raises(ValueError):
        export_leads_to_excel(
            records=[],
            evidence_by_website={},
            output_path=tmp_path / "empty.xlsx",
        )


def test_output_parent_directory_is_created(tmp_path: Path) -> None:
    """Missing output directories should be created automatically."""

    output_path = tmp_path / "nested" / "reports" / "leads.xlsx"

    result_path = export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": [make_evidence()]
        },
        output_path=output_path,
    )

    assert result_path.exists()
    assert output_path.parent.exists()


def test_workbook_file_is_created(tmp_path: Path) -> None:
    """The requested workbook file should be written."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": [make_evidence()]
        },
        output_path=output_path,
    )

    assert output_path.exists()


def test_workbook_contains_all_four_required_sheets(
    tmp_path: Path,
) -> None:
    """Workbook should include all required report sheets."""

    output_path = tmp_path / "leads.xlsx"
    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": [make_evidence()]
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)

    assert workbook.sheetnames == [
        "All Leads",
        "High Priority",
        "Evidence",
        "Run Summary",
    ]

    workbook.close()


def test_all_leads_has_correct_headers_and_data(
    tmp_path: Path,
) -> None:
    """The All Leads sheet should contain the configured columns and row data."""

    output_path = tmp_path / "leads.xlsx"
    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": [make_evidence()]
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["All Leads"]

    assert [
        cell.value
        for cell in worksheet[1]
    ] == LEAD_HEADERS
    assert worksheet["A2"].value == "Example City"
    assert worksheet["S2"].value == 55
    assert worksheet["T2"].value == "Medium"

    workbook.close()


def test_high_priority_contains_only_high_records(
    tmp_path: Path,
) -> None:
    """The High Priority sheet should filter out non-high records."""

    records = [
        make_record(
            name="High Example",
            website="https://high.example",
            priority="High",
            score=90,
        ),
        make_record(
            name="Medium Example",
            website="https://medium.example",
            priority="Medium",
            score=55,
        ),
    ]
    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=records,
        evidence_by_website={
            "https://high.example": [],
            "https://medium.example": [],
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["High Priority"]

    assert worksheet.max_row == 2
    assert worksheet["A2"].value == "High Example"

    workbook.close()


def test_high_priority_still_has_headers_when_empty(
    tmp_path: Path,
) -> None:
    """The High Priority sheet should exist even without high leads."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[
            make_record(
                priority="Medium",
                score=55,
            )
        ],
        evidence_by_website={
            "https://example-city.de": []
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["High Priority"]

    assert worksheet.max_row == 1
    assert [
        cell.value
        for cell in worksheet[1]
    ] == LEAD_HEADERS

    workbook.close()


def test_booleans_export_as_yes_no(tmp_path: Path) -> None:
    """Boolean fields should be readable Yes/No strings."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": []
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["All Leads"]

    assert worksheet["J2"].value == "Yes"
    assert worksheet["L2"].value == "No"

    workbook.close()


def test_lists_export_as_readable_joined_strings(
    tmp_path: Path,
) -> None:
    """List values should be joined with a readable separator."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[
            make_record(
                emails=[
                    "office@example-city.de",
                    "energy@example-city.de",
                ]
            )
        ],
        evidence_by_website={
            "https://example-city.de": []
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["All Leads"]

    assert worksheet["G2"].value == (
        "office@example-city.de; energy@example-city.de"
    )

    workbook.close()


def test_evidence_sheet_contains_all_evidence_rows(
    tmp_path: Path,
) -> None:
    """Every provided evidence item should be exported."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": [
                make_evidence(),
                SignalEvidence(
                    category="smart_city",
                    keyword="smart city",
                    excerpt="Fictional smart city evidence.",
                    source_url="https://example-city.de/smart",
                ),
            ]
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["Evidence"]

    assert worksheet.max_row == 3
    assert worksheet["A2"].value == "Example City"
    assert worksheet["D3"].value == "smart city"

    workbook.close()


def test_unknown_evidence_website_raises_value_error(
    tmp_path: Path,
) -> None:
    """Evidence for unknown websites should not be silently lost."""

    with pytest.raises(ValueError):
        export_leads_to_excel(
            records=[make_record()],
            evidence_by_website={
                "https://unknown.example": [make_evidence()]
            },
            output_path=tmp_path / "leads.xlsx",
        )


def test_run_summary_counts_priorities_correctly(
    tmp_path: Path,
) -> None:
    """Run Summary should count priority bands."""

    records = [
        make_record(
            name="High Example",
            website="https://high.example",
            priority="High",
            score=90,
        ),
        make_record(
            name="Medium Example",
            website="https://medium.example",
            priority="Medium",
            score=55,
        ),
        make_record(
            name="Low Example",
            website="https://low.example",
            priority="Low",
            score=15,
        ),
    ]
    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=records,
        evidence_by_website={
            "https://high.example": [],
            "https://medium.example": [],
            "https://low.example": [],
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["Run Summary"]

    assert worksheet["B2"].value == 1
    assert worksheet["B3"].value == 1
    assert worksheet["B4"].value == 1

    workbook.close()


def test_run_summary_totals_email_and_phone_counts(
    tmp_path: Path,
) -> None:
    """Run Summary should total exported email and phone counts."""

    records = [
        make_record(
            website="https://one.example",
            emails=["one@example.test", "two@example.test"],
            phone_numbers=["030 1111 2222"],
        ),
        make_record(
            website="https://two.example",
            emails=[],
            phone_numbers=["030 3333 4444", "030 5555 6666"],
        ),
    ]
    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=records,
        evidence_by_website={
            "https://one.example": [],
            "https://two.example": [],
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["Run Summary"]

    assert worksheet["B5"].value == 2
    assert worksheet["B6"].value == 3

    workbook.close()


def test_average_score_is_correct(tmp_path: Path) -> None:
    """Run Summary should include average lead score."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[
            make_record(
                website="https://one.example",
                score=100,
            ),
            make_record(
                website="https://two.example",
                score=50,
            ),
        ],
        evidence_by_website={
            "https://one.example": [],
            "https://two.example": [],
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["Run Summary"]

    assert worksheet["B8"].value == 75

    workbook.close()


def test_workbook_can_be_loaded_again_with_openpyxl(
    tmp_path: Path,
) -> None:
    """Saved workbook should be valid XLSX."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": []
        },
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert "All Leads" in workbook.sheetnames

    workbook.close()


def test_priority_cell_styling_differs_between_priority_levels(
    tmp_path: Path,
) -> None:
    """High, Medium, and Low priority cells should use different fills."""

    output_path = tmp_path / "leads.xlsx"
    records = [
        make_record(
            website="https://high.example",
            priority="High",
            score=90,
        ),
        make_record(
            website="https://medium.example",
            priority="Medium",
            score=55,
        ),
        make_record(
            website="https://low.example",
            priority="Low",
            score=15,
        ),
    ]

    export_leads_to_excel(
        records=records,
        evidence_by_website={
            "https://high.example": [],
            "https://medium.example": [],
            "https://low.example": [],
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)
    worksheet = workbook["All Leads"]
    fills = {
        worksheet["T2"].fill.fgColor.rgb,
        worksheet["T3"].fill.fgColor.rgb,
        worksheet["T4"].fill.fgColor.rgb,
    }

    assert len(fills) == 3

    workbook.close()


def test_top_row_is_frozen(tmp_path: Path) -> None:
    """Lead worksheets should freeze the header row."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": []
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)

    assert workbook["All Leads"].freeze_panes == "A2"
    assert workbook["High Priority"].freeze_panes == "A2"
    assert workbook["Evidence"].freeze_panes == "A2"

    workbook.close()


def test_auto_filter_is_configured(tmp_path: Path) -> None:
    """Tabular report sheets should expose an autofilter."""

    output_path = tmp_path / "leads.xlsx"

    export_leads_to_excel(
        records=[make_record()],
        evidence_by_website={
            "https://example-city.de": []
        },
        output_path=output_path,
    )

    workbook = load_export(output_path)

    assert workbook["All Leads"].auto_filter.ref == "A1:W2"
    assert workbook["High Priority"].auto_filter.ref == "A1:W1"
    assert workbook["Evidence"].auto_filter.ref == "A1:F1"

    workbook.close()
